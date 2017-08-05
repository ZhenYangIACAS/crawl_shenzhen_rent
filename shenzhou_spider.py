#-*- coding: UTF-8 -*-

import sys
import os
import time
import urllib
import urllib2
import codecs
import requests
import numpy as np
import random
from collections import OrderedDict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import * 

reload(sys)
sys.setdefaultencoding('utf8')

def html_file_analysis(storeHtml):
    
    spider_list=[]
    bs=BeautifulSoup(open(storeHtml), 'html.parser')
    for tag in bs.find_all("tbody"):
        for child_tag in tag.contents:
            #print("the child tag is ", child_tag)
            try:
                info=child_tag.find("td", class_="info")
                name=info.find("p", class_="name")
                other_info=info.find("p", class_="oth")

                ord_info=child_tag.find("td", class_="ord")
                pri_list = ord_info.select('em[class=num]')
                if len(pri_list)==1:
                    pri_list = ord_info.find_all("em", {'class':"num od_chagst"})+pri_list
                if len(pri_list)==0:
                    pri_list = ord_info.find_all("em", {'class':'num od_full'})
                
                assert(len(pri_list)==2)
                name=name.string.strip()
                other_info=other_info.string.strip()
                #print(name)

                price_per_day = pri_list[0].string.strip()
                price_total = pri_list[1].string.strip()
                #print(price_total)
                spider_list.append([name, other_info, price_per_day, price_total])

            except:
                continue
            #finally:
    return spider_list

def html_analysis(page_source):
    
    spider_list=[]
    if page_source is None:
        return None
        

    bs=BeautifulSoup(page_source, 'html.parser')
    for tag in bs.find_all("tbody"):
        for child_tag in tag.contents:
            #print("the child tag is ", child_tag)
            try:
                info=child_tag.find("td", class_="info")
                name=info.find("p", class_="name")
                other_info=info.find("p", class_="oth")

                ord_info=child_tag.find("td", class_="ord")
                pri_list = ord_info.select('em[class=num]')
                if len(pri_list)==1:
                    pri_list = ord_info.find_all("em", {'class':"num od_chagst"})+pri_list
                if len(pri_list)==0:
                    pri_list = ord_info.find_all("em", {'class':'num od_full'})
                
                assert(len(pri_list)==2)
                name=name.string.strip()
                other_info=other_info.string.strip()
                #print(name)

                price_per_day = pri_list[0].string.strip()
                price_total = pri_list[1].string.strip()
                #print(price_total)
                spider_list.append([name, other_info, price_per_day, price_total])

            except:
                continue
            #finally:
    return spider_list

def shenzhou_spider_init():
    browser=webdriver.Firefox()
    url='https://order.zuche.com'
    browser.get(url)
    time.sleep(10)

    browser.find_element_by_link_text('到店取还').click()
    time.sleep(10)
    return browser

def shenzhou_spider(browser_init, cityFrom, cityTo, zoneFrom, storeFrom, zoneTo, storeTo):

    
    browser=browser_init
    try:
        fromCity=browser.find_element_by_id("fromCity")
        fromCity.clear()
        #fromCity.send_keys(u'北京')
        fromCity.send_keys(cityFrom)

        time.sleep(4)
        fromStore=browser.find_element_by_id("fromStoreName")
        fromStore.click()
        time.sleep(6)
        fromZone=browser.find_element_by_partial_link_text(zoneFrom)
        #fromZone=browser.find_element_by_link_text("机场/火车站")
        chain=ActionChains(browser)
        chain.move_to_element(fromZone).perform()
        time.sleep(4)
        
        #browser.find_element_by_link_text("北京站店").click()
        browser.find_element_by_partial_link_text(storeFrom).click()
    
    
        toCity =browser.find_element_by_id("toCityName")
        toCity.clear()
        toCity.send_keys(cityTo)

        time.sleep(4)
        toStoreName=browser.find_element_by_id("toStoreName")
        toStoreName.click()
        time.sleep(4)
        toZone=browser.find_element_by_partial_link_text(zoneTo)
        #toZone=browser.find_element_by_link_text("机场/火车站")
        chain_sec=ActionChains(browser)
        chain_sec.move_to_element(toZone).perform()
        time.sleep(4)

        browser.find_element_by_partial_link_text(storeTo).click()
        #browser.find_element_by_link_text("北京站店").click()

        time.sleep(2)
        browser.find_element_by_link_text("立即选车").click()
    
        time.sleep(20)

        show_full = browser.find_element_by_class_name("show_odfull")
        show_full.click()
        browser.get_screenshot_as_file("./screen.png")
        page_source=browser.page_source
        pageFile=open(cityFrom.replace('/', '')+'_'+zoneFrom.replace('/', '')+'_'+storeFrom.replace('/', '')+'.html', 'w')
        pageFile.write(page_source)
        pageFile.close()
    except:
        page_source=None
    
    finally:
        browser.quit()

    return page_source

def read_zone_store_info(zone_store_file):
    """the template of the file is:
       火车站
       北京南站  北京西站 北京站
    """
    zone_store_dic=OrderedDict()

    zone_store_list=[]
    inFile=open(zone_store_file, 'r')
    for line in inFile.readlines():
        line=line.strip()
        line_list=line.split()
        zone_store_list.append(line_list)

    assert(len(zone_store_list) %2 ==0)
    for i in range(0, len(zone_store_list),2):
        zone_store_dic[zone_store_list[i][0]]=zone_store_list[i+1]

    return zone_store_dic

def write_to_excel(sheet_list, spider_lists, save_path):

    wb=Workbook()
    ws=[]

    assert(len(sheet_list)==len(spider_lists))

    for i in range(len(sheet_list)):
        ws.append(wb.create_sheet(title=sheet_list[i].decode()))

    for i in range(len(spider_lists)):
        ws[i].append(['number', 'name', 'info', 'price_per_day', 'price_total'])
        count=1
        if spider_lists[i] is None:
            continue

        for bl in spider_lists[i]:
            if bl is None:
                ws[i].append([count, 'xx', 'xx', 0., 0.])
            else:
                ws[i].append([count, bl[0], bl[1], float(bl[2]), float(bl[3])])
            count+=1

    wb.save(save_path)

def get_spider_and_sheet_list_from_dir(file_dir):
    spider_lists=[]
    sheet_lists=[]
    dir_list=os.listdir(file_dir)
    for i in range(0, len(dir_list)):
        path=os.path.join(file_dir, dir_list[i])
        if os.path.isfile(path):
            #print(os.path.basename(path))
            sheet_name=os.path.basename(path).split('.')[0]
            sheet_lists.append(sheet_name)
            spider_list=html_file_analysis(path)
            spider_lists.append(spider_list)
    return spider_lists, sheet_lists

if __name__=='__main__':
    cityFrom=cityTo=u'上海'
    zone_store_dic=read_zone_store_info(sys.argv[1])
    
    spider_lists=[]
    sheet_list=[]

    #file_dir="spider_file"
    #spider_lists, sheet_lists=get_spider_and_sheet_list_from_dir(file_dir)
    #write_to_excel(sheet_lists, spider_lists, cityFrom+"_spider_result.xlsx")

    
    for key, value in zone_store_dic.items():
        for v in value:
            key=key.strip()
            v=v.strip()
            sheet_list.append(cityFrom.replace('/', '')+'-'+key.replace('/', '')+'-'+v.replace('/', ''))
            
            try:
                browser_init=shenzhou_spider_init()
                page_source =shenzhou_spider(browser_init, cityFrom, cityTo, key, v, key, v)
                spider_list=html_analysis(page_source)
            except:
                spider_list=None
                print("can not crawl the web page the %s store of %s zone or analysis error" %(key, v))

            finally:
                spider_lists.append(spider_list)

    write_to_excel(sheet_list, spider_lists, "spider_result.xlsx")




